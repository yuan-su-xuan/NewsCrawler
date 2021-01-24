import xlrd
import xlwt
from bs4 import BeautifulSoup
import time
import Crawler
import openpyxl
from openpyxl import Workbook
workbook = Workbook()


def getNews(xlsPath):
    readbook = xlrd.open_workbook(xlsPath)
    for i in range(0, 4):
        #对一个xls文件的不同sheet读取链接获取新闻内容
        sheet = readbook.sheet_by_index(i)
        writeText(i, sheet)
        workbook.save('新闻内容.xlsx')
        time.sleep(60)

def writeText(i, sheet):
    #将链接中的新闻内容写入文件中
    toWriteSheet = workbook.create_sheet("第" + str(i) + "阶段")
    nrows = sheet.nrows
    row = 0
    for i in range(1, nrows):
        src=sheet.cell(i,0).value
        #过滤不需要数据
        if 'chinairn' in src or '凤凰' in src:
            print('跳过无效url')
            continue
        name = sheet.cell(i, 1).value
        url = sheet.cell(i, 2).value
        newsTime = sheet.cell(i,3).value
        print('获得表中第'+str(i)+'行')
        #获得新闻内容
        text = getText(url)
        if text is None:
            print('跳过无效url')
            continue
        arow=[src,name,text,newsTime]
        toWriteSheet.append(arow)
        print('第'+str(row)+'行已被读入')
        row += 1
        workbook.save('新闻内容.xlsx')



def getText(url):
    #访问链接并获得新闻内容
    html = Crawler.askUrl(url)
    if html == None:
        return None
    soup = BeautifulSoup(html, "html.parser")
    contents = soup.find_all('span', {'class': 'bjh-p'})
    if len(contents) == 0:
        contents = soup.find_all('p')
    text = ''
    for i in range(0, len(contents)):
        text += contents[i].text
    return text
